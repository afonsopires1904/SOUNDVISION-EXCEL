"""
Soundvision PDF Extractor — Generic
Reads all PDFs from ../data/ and outputs Excel files to ../output/

Usage:
    python extract.py                  # processes all PDFs in ../data/
    python extract.py report.pdf       # processes a specific file in ../data/
"""

import sys
import re
import pdfplumber
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

SRC_DIR    = Path(__file__).parent
DATA_DIR   = SRC_DIR.parent / "data"
OUTPUT_DIR = SRC_DIR.parent / "output"

def extract_text(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def extract_metadata(text):
    """Extract report title (file name) and date from the Soundvision header."""
    import re
    name_match = re.search(r"File name:\s*(.+)", text)
    date_match = re.search(r"Date:\s*(\d{4}/\d{2}/\d{2})", text)
    name = name_match.group(1).strip().replace(".xmlp", "") if name_match else "Unknown"
    date = date_match.group(1).strip() if date_match else ""
    return name, date


def get_canonical_name(source_name):
    name = re.sub(r'\s+[LR]\s*(\d*)$', lambda m: (' ' + m.group(1)) if m.group(1) else '', source_name).strip()
    return name

def parse_physical_config(block):
    fields = {
        "Configuration":     r"Configuration:\s*(.+)",
        "Bumper":            r"Bumper:\s*(.+)",
        "# Motors":          r"# motors:\s*(\d+)",
        "Position X (m)":    r"Position \(X; Y; Z, m\):\s*([\-\d.]+);",
        "Position Y (m)":    r"Position \(X; Y; Z, m\):\s*[\-\d.]+;\s*([\-\d.]+);",
        "Position Z (m)":    r"Position \(X; Y; Z, m\):\s*[\-\d.]+;\s*[\-\d.]+;\s*([\-\d.]+)",
        "Site (°)":          r"Site:\s*([\-\d.]+)\s*°",
        "Azimuth (°)":       r"Azimuth:\s*([\-\d.]+)\s*°",
        "Bottom Elev. (m)":  r"Bottom elevation:\s*([\-\d.]+)",
        "Top Site (°)":      r"Top site:\s*([\-\d.]+)\s*°",
        "Bottom Site (°)":   r"Bottom site:\s*([\-\d.]+)\s*°",
        "Total Weight (kg)": r"Total weight \(Enclosures \+ Frames\):\s*([\d.]+)",
        "Enclosure Wt (kg)": r"Total enclosure weight:\s*([\d.]+)",
        "Front Motor (kg)":  r"Front motor load:\s*([\d.]+)",
        "Rear Motor (kg)":   r"Rear motor load:\s*([\d.]+)",
        "Front L Motor (kg)":r"Front left motor load:\s*([\d.]+)",
        "Rigging Points":    r"(?:Front|Rear).*?pickup position.*?:\s*(EXT|INT)\s*\(",
        "Front R Motor (kg)":r"Front right motor load:\s*([\d.]+)",
    }
    config = {}
    for key, pattern in fields.items():
        m = re.search(pattern, block)
        if m:
            config[key] = m.group(1).strip()
    return config

def parse_enclosure_table(block):
    has_panflex = bool(re.search(r"Panflex", block))
    has_angles  = bool(re.search(r"Angles \(°\)", block))

    if has_angles:
        # Token-based parser: handles K1 (no Panflex), K2 (with Panflex),
        # and mixed arrays. Type = all leading alpha tokens; numbers follow.
        line_pat = re.compile(r"^#(\d+)\s+(.+)$", re.MULTILINE)
        num_pat  = re.compile(r"^-?\d+(?:\.\d+)?$")
        pflex_pat = re.compile(r"^\d+/\d+$")
        rows = []
        for m in line_pat.finditer(block):
            enc_num = int(m.group(1))
            tokens  = m.group(2).split()
            # Split into type tokens (start with letter) and value tokens (numbers/panflex)
            type_tokens = []
            val_tokens  = []
            for t in tokens:
                if not type_tokens and not num_pat.match(t) and not pflex_pat.match(t):
                    type_tokens.append(t)
                elif type_tokens and not num_pat.match(t) and not pflex_pat.match(t):
                    type_tokens.append(t)
                else:
                    val_tokens.append(t)
            if not type_tokens or not val_tokens:
                continue
            cab_type = " ".join(type_tokens)
            # Panflex is last token if it matches NN/NN pattern
            panflex = "—"
            nums = []
            for t in val_tokens:
                if pflex_pat.match(t):
                    panflex = t
                elif num_pat.match(t):
                    nums.append(float(t))
            if not nums:
                continue
            angle = nums[0]
            parts = panflex.split("/") if "/" in panflex else [panflex, panflex]
            rows.append({
                "Enc #":     enc_num,
                "Type":      cab_type,
                "Angle (°)": angle,
                "Circuit":   "",
                "Panflex L": parts[0],
                "Panflex R": parts[1] if len(parts) > 1 else parts[0],
                "Amp ID L":  "",
                "Amp ID R":  "",
                "Amp Ch":    "",
            })
        rows.sort(key=lambda r: r["Enc #"])
        return rows, ["Enc #", "Type", "Angle (°)", "Circuit", "Panflex L", "Panflex R", "Amp ID L", "Amp ID R", "Amp Ch"]

    else:
        # Subs / point source (KS28, SB28, X8...)
        pattern = re.compile(
            r"^#(\d+)\s+([\w]+(?:_C)?)\s+[-\d.]+\s+[-\d.]+\s+[-\d.]+\s*$",
            re.MULTILINE
        )
        rows = []
        for m in pattern.finditer(block):
            rows.append({
                "Enc #":     int(m.group(1)),
                "Type":      m.group(2).strip(),
                "Circuit":   "",
                "Amp ID L":  "",
                "Amp ID R":  "",
                "Amp Ch":    "",
            })
        return rows, ["Enc #", "Type", "Circuit", "Amp ID L", "Amp ID R", "Amp Ch"]

def split_source_blocks(text):
    pattern = re.compile(r"^\d+\.\s+Source:\s+(.+)$", re.MULTILINE)
    matches = list(pattern.finditer(text))
    blocks = []
    for i, m in enumerate(matches):
        start = m.start()
        end   = matches[i + 1].start() if i + 1 < len(matches) else len(text)
        blocks.append((m.group(1).strip(), text[start:end]))
    return blocks

def get_group_for_source(text, source_name):
    pattern = re.compile(rf"\d+\.\s+Source:\s+{re.escape(source_name)}")
    m = pattern.search(text)
    if not m:
        return "Unknown"
    group_pattern = re.compile(r"^\d+\.\s+Group:\s+(.+)$", re.MULTILINE)
    groups_before = list(group_pattern.finditer(text, 0, m.start()))
    if not groups_before:
        return "Unknown"
    for g in reversed(groups_before):
        name = g.group(1).strip()
        if name.upper() != "ALL":
            return name
    return "Unknown"

def source_fingerprint(physical, enclosures):
    normalised = {}
    for k, v in physical.items():
        if k in ("Position X (m)", "Azimuth (°)"):
            try:
                v = str(abs(float(v)))
            except ValueError:
                pass
        normalised[k] = v
    phys_tuple = tuple(sorted(normalised.items()))
    from collections import Counter
    enc_tuple = tuple(sorted(Counter(e["Type"] for e in enclosures).items()))
    return (phys_tuple, enc_tuple)

def parse_document(text):
    source_blocks = split_source_blocks(text)
    groups = {}
    for source_name, block in source_blocks:
        group            = get_group_for_source(text, source_name)
        physical         = parse_physical_config(block)
        enclosures, cols = parse_enclosure_table(block)
        if "stacked" in physical.get("Configuration", "").lower():
            continue
        fp = source_fingerprint(physical, enclosures)
        if any(source_fingerprint(s["physical"], s["enclosures"]) == fp
               for s in groups.get(group, [])):
            continue
        canonical = get_canonical_name(source_name)
        entry = {
            "name":       canonical,
            "physical":   physical,
            "enclosures": enclosures,
            "columns":    cols,
        }
        if group not in groups:
            groups[group] = []
        groups[group].append(entry)
    return groups

def enclosure_group_size(enclosures):
    """Returns the colour-group size based on speaker model."""
    if not enclosures:
        return 1
    model = enclosures[0].get("Type", "").upper()
    if any(x in model for x in ("K1", "K2", "K3")):
        return 4
    return 3  # KARA, KIVA and everything else

def enc_color_index(idx, enclosures):
    """
    Colour index that groups by cabinet type size (K1/K2/K3 = 4, rest = 3).
    When the type changes, flips only if the previous group ended mid-cycle
    (if it ended on a boundary the colour already flipped naturally).
    """
    color = 0
    count = 0
    prev_type = None

    for i, enc in enumerate(enclosures[:idx + 1]):
        model = enc.get("Type", "").upper()
        size  = 4 if any(x in model for x in ("K1", "K2", "K3")) else 3

        if model != prev_type:
            if prev_type is not None:
                # Use the PREVIOUS type's group size to check if mid-cycle
                prev_size = 4 if any(x in prev_type for x in ("K1", "K2", "K3")) else 3
                if count % prev_size != 0:
                    color = 1 - color
            count = 0
            prev_type = model

        if i < idx:
            count += 1
            if count % size == 0:
                color = 1 - color

    return color

def thin_border():
    s = Side(style="thin", color="B0B0B0")
    return Border(left=s, right=s, top=s, bottom=s)

HEADER_FILL  = PatternFill("solid", start_color="282C34")
SUBHEAD_FILL = PatternFill("solid", start_color="3A3F4B")
ALT_FILL     = PatternFill("solid", start_color="FAF3E0")
WHITE_FILL   = PatternFill("solid", start_color="FAFAFA")
ROW_FILL     = PatternFill("solid", start_color="282C34")
CARD_FILL    = PatternFill("solid", start_color="F9DCDC")
INPUT_FILL   = PatternFill("solid", start_color="FFF9E6")  # light yellow for user-fillable cells
SECTION_FILL = PatternFill("solid", start_color="E8E0D0")  # warm grey for section labels
HEADER_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=13)
SUBHEAD_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=10)
LABEL_FONT   = Font(name="Arial", bold=True, color="3A3F4B", size=10)
BODY_FONT    = Font(name="Arial", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT         = Alignment(horizontal="left",   vertical="center")

# ── Field grouping for physical config display ───────────────────────────────

FIELD_GROUPS = [
    ("Rigging",     ["Configuration", "Bumper", "# Motors", "Rigging Points", "Total Weight (kg)", "Front Motor (kg)", "Rear Motor (kg)", "Front L Motor (kg)", "Front R Motor (kg)"]),
    ("Position",    ["Position X (m)", "Position Y (m)"]),
    ("Orientation", ["Site (°)", "Azimuth (°)"]),
    ("Elevation",   ["Position Z (m)", "Bottom Elev. (m)", "Bottom Site (°)"]),
]

def grouped_physical_items(physical):
    """
    Returns physical config fields ordered and grouped by FIELD_GROUPS.
    Each group is a (group_label, [(key, val), ...]) tuple.
    Only includes fields that exist in the physical dict.
    """
    result = []
    for group_label, keys in FIELD_GROUPS:
        items = [(k, physical[k]) for k in keys if k in physical]
        if items:
            result.append((group_label, items))
    return result


CARD_FILL = PatternFill("solid", start_color="F4CCCC")  # soft red for cardioid cabinets

# Resistor colour code palette — vivid colours matching resistor bands
CIRCUIT_COLORS = {
    "A": "8B4513",  # Brown  (1)
    "B": "FF0000",  # Red    (2)
    "C": "FF6600",  # Orange (3)
    "D": "FFD700",  # Yellow (4)
    "E": "008000",  # Green  (5)
    "F": "0000FF",  # Blue   (6)
    "G": "8B00FF",  # Violet (7)
    "H": "808080",  # Grey   (8)
    "I": "F0F0F0",  # White  (9)
    "J": "222222",  # Black  (0)
}

def _circuit_fill(val):
    hex_color = CIRCUIT_COLORS.get(str(val).strip().upper())
    if hex_color:
        return PatternFill("solid", start_color=hex_color)
    return None

def write_excel(groups, output_path, report_name="", report_date=""):
    wb = Workbook()
    wb.remove(wb.active)

    # ── Cover sheet ───────────────────────────────────────────────────────────
    cover = wb.create_sheet(title="Report Info", index=0)
    cover.sheet_view.showGridLines = False

    cover.merge_cells("A1:G1")
    c = cover["A1"]
    c.value = report_name or "Soundvision Report"
    c.font = Font(name="Arial", bold=True, color="F5A623", size=20)
    c.fill = PatternFill("solid", start_color="282C34")
    c.alignment = Alignment(horizontal="left", vertical="center")
    cover.row_dimensions[1].height = 48

    cover.merge_cells("A2:G2")
    c = cover["A2"]
    c.value = f"Date: {report_date}" if report_date else ""
    c.font = Font(name="Arial", color="AAAAAA", size=11)
    c.fill = PatternFill("solid", start_color="282C34")
    c.alignment = Alignment(horizontal="left", vertical="center")
    cover.row_dimensions[2].height = 24

    cover.merge_cells("A3:G3")
    cover["A3"].fill = PatternFill("solid", start_color="282C34")
    cover.row_dimensions[3].height = 10

    # ── User input fields ─────────────────────────────────────────────────────
    input_fields = [
        ("System Engineer:", 4),
        ("Company:",         5),
        ("Venue:",           6),
        ("Date:",            7),
    ]
    for label, r in input_fields:
        lc = cover.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", bold=True, color="3A3F4B", size=10)
        lc.fill = PatternFill("solid", start_color="E8E0D0")
        lc.alignment = Alignment(horizontal="left", vertical="center")
        lc.border = thin_border()
        cover.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        s_thin = Side(style="thin", color="B0B0B0")
        s_black = Side(style="thin", color="000000")
        for ci in range(2, 6):
            cell = cover.cell(row=r, column=ci)
            cell.fill = PatternFill("solid", start_color="FFF9E6")
            cell.alignment = Alignment(horizontal="left", vertical="center")
            left  = s_black if ci == 2 else Side(border_style=None)
            right = s_black if ci == 5 else Side(border_style=None)
            cell.border = Border(left=left, right=right, top=s_thin, bottom=s_thin)
        cover.cell(row=r, column=2).value = ""
        cover.cell(row=r, column=2).font = Font(name="Arial", size=10)
        cover.row_dimensions[r].height = 18

    # ── Summary section header ────────────────────────────────────────────────
    cover.row_dimensions[8].height = 8
    cover.merge_cells("A9:I9")
    sh = cover["A9"]
    sh.value = "System Summary"
    sh.font = Font(name="Arial", bold=True, color="F5A623", size=11)
    sh.fill = PatternFill("solid", start_color="282C34")
    sh.alignment = Alignment(horizontal="left", vertical="center")
    cover.row_dimensions[9].height = 22

    for col, width in zip("ABCDEFGHIJK", [18, 14, 10, 10, 10, 8, 9, 9, 9, 9, 9]):
        cover.column_dimensions[col].width = width
    cover.column_dimensions["L"].width = 3   # spacer
    cover.column_dimensions["M"].width = 3   # spacer
    cover.column_dimensions["N"].width = 3   # spacer
    cover.column_dimensions["O"].width = 3   # spacer
    cover.column_dimensions["P"].width = 55  # instructions

    # ── Instructions panel (column P, far from data) ──────────────────────────
    instr_title = cover.cell(row=1, column=16, value="📖 How to use this file")
    instr_title.font = Font(name="Arial", bold=True, color="F5A623", size=11)
    instr_title.fill = PatternFill("solid", start_color="282C34")
    instr_title.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cover.row_dimensions[1].height = 48

    instructions = [
        ("Step 1 — Review the data", "The Soundvision report has been automatically parsed. Each group (MAIN LR, OUTFILL, etc.) has its own sheet with physical config and enclosure geometry."),
        ("Step 2 — Fill in your details", "Enter System Engineer, Company, Venue and Date in the yellow fields on the left."),
        ("Step 3 — Assign circuits", "In the System Summary table below, use the Circuit column dropdown (A–J) to assign each enclosure to an amplifier circuit.\n\nColour code: A=Brown  B=Red  C=Orange  D=Yellow  E=Green  F=Blue  G=Violet  H=Grey  I=White  J=Black"),
        ("Step 4 — Fill Amp IDs and channels", "Enter Amp ID L, Amp ID R and Amp Ch for each enclosure. These update automatically on the individual group sheets."),
        ("About the PDF report", "The PDF is a rigging reference — it contains physical configuration and enclosure geometry only. Circuit assignments and amp data are Excel-only, as they are filled in after the fact."),
        ("Note — Screenshots", "The data table ends at column K. This instructions panel is intentionally placed here so it does not appear in screenshots of the data."),
    ]

    instr_row = 2
    for title, body in instructions:
        t = cover.cell(row=instr_row, column=16, value=title)
        t.font = Font(name="Arial", bold=True, color="3A3F4B", size=9)
        t.fill = PatternFill("solid", start_color="E8E0D0")
        t.alignment = Alignment(horizontal="left", vertical="center")
        t.border = thin_border()
        cover.row_dimensions[instr_row].height = 14
        instr_row += 1

        b = cover.cell(row=instr_row, column=16, value=body)
        b.font = Font(name="Arial", size=9)
        b.fill = WHITE_FILL
        b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        b.border = thin_border()
        cover.row_dimensions[instr_row].height = 70 if "Colour code" in body else 45
        instr_row += 1

    # ── Group sheets (populate _enc_start_rows first) ─────────────────────────
    _enc_start_rows = {}
    for group_name, sources in groups.items():
        sheet_name = re.sub(r'[\\/*?:\[\]]', '', group_name)[:31]
        ws = wb.create_sheet(title=sheet_name)
        ws.sheet_view.showGridLines = False
        # Physical config label columns (A, D) and value columns (B+C, E+F)
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 4
        ws.column_dimensions["D"].width = 22
        ws.column_dimensions["E"].width = 20
        ws.column_dimensions["F"].width = 4
        row = 1
        # Report name on its own row above the group title
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = report_name
        c.font = Font(name="Arial", color="777777", size=9, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 14
        row += 1
        ws.merge_cells(f"A{row}:G{row}")
        c = ws[f"A{row}"]
        c.value = f"Group: {group_name}"
        c.font = Font(name="Arial", bold=True, color="F5A623", size=14)
        c.fill = HEADER_FILL; c.alignment = CENTER
        ws.row_dimensions[row].height = 30
        row += 2
        for source in sources:
            ws.merge_cells(f"A{row}:G{row}")
            c = ws[f"A{row}"]
            c.value = source["name"]
            c.font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            c.fill = SUBHEAD_FILL; c.alignment = LEFT
            ws.row_dimensions[row].height = 22
            row += 1
            physical = source["physical"]
            if physical:
                dark_fill = PatternFill("solid", start_color="4A5060")
                ws.merge_cells(f"A{row}:G{row}")
                for col_idx in range(1, 8):
                    ws.cell(row=row, column=col_idx).fill = dark_fill
                c = ws[f"A{row}"]
                c.value = "Physical Configuration"
                c.font = Font(name="Arial", bold=True, color="F5A623", size=10)
                c.alignment = LEFT
                ws.row_dimensions[row].height = 18
                row += 1
                for group_label, items in grouped_physical_items(physical):
                    # Group sub-header
                    ws.merge_cells(f"A{row}:G{row}")
                    c = ws[f"A{row}"]
                    c.value = group_label
                    c.font = Font(name="Arial", bold=True, color="F5A623", size=9)
                    c.fill = SECTION_FILL; c.alignment = LEFT
                    ws.row_dimensions[row].height = 13
                    row += 1
                    for i in range(0, len(items), 2):
                        fill = ALT_FILL if (row % 2 == 0) else WHITE_FILL
                        for offset, idx in enumerate([i, i + 1]):
                            if idx >= len(items):
                                break
                            key, val = items[idx]
                            col_label = offset * 3 + 1
                            lc = ws.cell(row=row, column=col_label, value=key)
                            lc.font = LABEL_FONT; lc.alignment = LEFT
                            lc.fill = fill; lc.border = thin_border()
                            vc = ws.cell(row=row, column=col_label + 1, value=val)
                            vc.font = BODY_FONT; vc.alignment = LEFT
                            vc.fill = fill; vc.border = thin_border()
                            ws.merge_cells(start_row=row, start_column=col_label + 1,
                                           end_row=row, end_column=col_label + 2)
                        row += 1
                row += 1
            enclosures = source["enclosures"]
            columns    = source["columns"]
            if enclosures:
                dark_fill = PatternFill("solid", start_color="4A5060")
                ws.merge_cells(f"A{row}:G{row}")
                for col_idx in range(1, 8):
                    ws.cell(row=row, column=col_idx).fill = dark_fill
                c = ws[f"A{row}"]
                c.value = "Per-Enclosure Geometry"
                c.font = Font(name="Arial", bold=True, color="F5A623", size=10)
                c.alignment = LEFT
                ws.row_dimensions[row].height = 15
                row += 1
                # Two-row header: group row + L/R subheader row
                group_headers = {
                    "Panflex L": "Panflex", "Panflex R": "Panflex",
                    "Amp ID L":  "Amp ID",  "Amp ID R":  "Amp ID",
                }
                # Row 1: group labels (merged for L/R pairs)
                written_groups = set()
                for col, h in enumerate(columns, 1):
                    group = group_headers.get(h)
                    if group and group not in written_groups:
                        # Find the R column
                        r_key = h.replace(" L", " R")
                        r_col = columns.index(r_key) + 1 if r_key in columns else col
                        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=r_col)
                        c = ws.cell(row=row, column=col, value=group)
                        c.font = Font(name="Arial", bold=True, color="F5A623", size=9)
                        c.fill = ROW_FILL; c.alignment = CENTER; c.border = thin_border()
                        written_groups.add(group)
                    elif not group:
                        # Span two rows for non-grouped columns
                        ws.merge_cells(start_row=row, start_column=col, end_row=row+1, end_column=col)
                        c = ws.cell(row=row, column=col, value=h)
                        c.font = Font(name="Arial", bold=True, color="F5A623", size=9)
                        c.fill = ROW_FILL; c.alignment = CENTER; c.border = thin_border()
                ws.row_dimensions[row].height = 13
                row += 1
                # Row 2: L / R subheaders
                for col, h in enumerate(columns, 1):
                    if h in group_headers:
                        label = "L" if h.endswith(" L") else "R"
                        c = ws.cell(row=row, column=col, value=label)
                        c.font = Font(name="Arial", bold=True, color="F5A623", size=9)
                        c.fill = ROW_FILL; c.alignment = CENTER; c.border = thin_border()
                    else:
                        # Already merged from row above, just set border
                        ws.cell(row=row, column=col).border = thin_border()
                ws.row_dimensions[row].height = 13
                row += 1
                _enc_start_rows[f"{sheet_name}__{source['name']}"] = row
                for idx, enc in enumerate(enclosures):
                    is_cardioid = enc.get("Type", "").endswith("_C")
                    if is_cardioid:
                        row_fill = CARD_FILL
                    else:
                        row_fill = ALT_FILL if enc_color_index(idx, enclosures) == 0 else WHITE_FILL
                    for col, key in enumerate(columns, 1):
                        val = enc.get(key, "")
                        if key == "Circuit":
                            cfill = _circuit_fill(val)
                            fill = cfill if cfill else INPUT_FILL
                        elif key in ("Amp ID L", "Amp ID R", "Amp Ch"):
                            fill = INPUT_FILL
                        else:
                            fill = row_fill
                        c = ws.cell(row=row, column=col, value=val)
                        c.font = BODY_FONT; c.alignment = CENTER
                        c.fill = fill; c.border = thin_border()
                    row += 1
            # Dropdown validation for Circuit column
            if "Circuit" in columns:
                circuit_col_idx = columns.index("Circuit") + 1
                circuit_col_letter = get_column_letter(circuit_col_idx)
                enc_start = row - len(enclosures)
                enc_end   = row - 1
                dv = DataValidation(
                    type="list",
                    formula1='"A,B,C,D,E,F,G,H,I,J"',
                    allow_blank=True,
                )
                ws.add_data_validation(dv)
                dv.add(f"{circuit_col_letter}{enc_start}:{circuit_col_letter}{enc_end}")

                # Conditional formatting — cell background by circuit letter
                circuit_range = f"{circuit_col_letter}{enc_start}:{circuit_col_letter}{enc_end}"
                for letter, hex_color in CIRCUIT_COLORS.items():
                    ws.conditional_formatting.add(circuit_range, CellIsRule(
                        operator="equal",
                        formula=[f'"{letter}"'],
                        fill=PatternFill("solid", start_color=hex_color, end_color=hex_color)
                    ))

            from openpyxl.worksheet.pagebreak import Break
            ws.row_breaks.append(Break(id=row))
            row += 2
        # Col map: Enc#, Type, Angle, PanflexL, PanflexR, AmpIDL, AmpIDR, AmpCh
        # Enclosure table columns — only set G, H, I (Amp cols)
        # A-F keep the physical config widths set above
        for col, width in [("G", 9), ("H", 9), ("I", 9)]:
            ws.column_dimensions[col].width = width
        # Also fix column C width for Angle on group sheets
        ws.column_dimensions["C"].width = 10

    # ── Fill cover summary — user-input cols editable, rest static ─────────────
    # Columns in summary: Group(1), Source(2), Enc#(3), Type(4), Angle(5),
    #   Circuit(6), Panflex L(7), Panflex R(8), Amp ID L(9), Amp ID R(10), Amp Ch(11)
    USER_INPUT_COLS = {6, 9, 10, 11}  # Circuit, Amp ID L, Amp ID R, Amp Ch
    GROUP_COLORS = [
        "D6E4F0",  # soft blue
        "D5F0E0",  # soft green
        "E8D5F0",  # soft purple
        "D5EEF0",  # soft teal
        "F0D5EC",  # soft pink
        "D5E8F0",  # soft sky
        "E0F0D5",  # soft lime
        "F0E8D5",  # soft peach
    ]
    sum_headers = ["Group", "Source", "Enc #", "Type", "Angle (°)",
                   "Circuit", "Panflex L", "Panflex R", "Amp ID L", "Amp ID R", "Amp Ch"]
    _cover_enc_rows = {}  # key: "sheet__source__enc_idx" -> cover row

    sum_row = 11
    for g_idx, (group_name, sources) in enumerate(groups.items()):
        sheet_name = re.sub(r'[\/*?:\[\]]', '', group_name)[:31]
        group_color = PatternFill("solid", start_color=GROUP_COLORS[g_idx % len(GROUP_COLORS)])

        # Group banner
        cover.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=11)
        gh = cover.cell(row=sum_row, column=1, value=f"Group: {group_name}")
        gh.font = Font(name="Arial", bold=True, color="F5A623", size=9)
        gh.fill = PatternFill("solid", start_color="282C34")
        gh.alignment = Alignment(horizontal="left", vertical="center")
        for ci in range(1, 12):
            cover.cell(row=sum_row, column=ci).border = thin_border()
        cover.row_dimensions[sum_row].height = 14
        sum_row += 1

        for source in sources:
            # Source sub-header
            cover.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=11)
            sh = cover.cell(row=sum_row, column=1, value=source["name"])
            sh.font = Font(name="Arial", bold=True, color="FFFFFF", size=8)
            sh.fill = PatternFill("solid", start_color="3A3F4B")
            sh.alignment = Alignment(horizontal="left", vertical="center")
            for ci in range(1, 12): cover.cell(row=sum_row, column=ci).border = thin_border()
            cover.row_dimensions[sum_row].height = 12
            sum_row += 1

            # Physical config compact
            phys_items = list(source["physical"].items())
            for i in range(0, len(phys_items), 2):
                fill = ALT_FILL if (sum_row % 2 == 0) else WHITE_FILL
                for offset, idx2 in enumerate([i, i+1]):
                    if idx2 >= len(phys_items): break
                    k, v = phys_items[idx2]
                    col_l = offset * 4 + 1
                    lc = cover.cell(row=sum_row, column=col_l, value=k)
                    lc.font = Font(name="Arial", bold=True, color="3A3F4B", size=7)
                    lc.fill = fill; lc.alignment = Alignment(horizontal="left", vertical="center")
                    lc.border = thin_border()
                    vc = cover.cell(row=sum_row, column=col_l+1, value=str(v))
                    vc.font = Font(name="Arial", size=7)
                    vc.fill = fill; vc.alignment = Alignment(horizontal="left", vertical="center")
                    vc.border = thin_border()
                    cover.merge_cells(start_row=sum_row, start_column=col_l+1,
                                      end_row=sum_row, end_column=col_l+2)
                cover.row_dimensions[sum_row].height = 11
                sum_row += 1
            sum_row += 1  # spacer after physical config

            enclosures = source["enclosures"]
            if not enclosures:
                continue

            # Column headers
            for col_idx, h in enumerate(sum_headers, 1):
                c = cover.cell(row=sum_row, column=col_idx, value=h)
                c.font = Font(name="Arial", bold=True, color="F5A623", size=9)
                c.fill = PatternFill("solid", start_color="282C34")
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()
            cover.row_dimensions[sum_row].height = 14
            sum_row += 1

            # Enclosure rows
            for enc_idx, enc in enumerate(enclosures):
                is_cardioid = enc.get("Type", "").endswith("_C")
                row_fill = CARD_FILL if is_cardioid else (
                    ALT_FILL if enc_color_index(enc_idx, enclosures) == 0 else WHITE_FILL
                )
                static_vals = [
                    group_name,
                    source["name"],
                    enc.get("Enc #", ""),
                    enc.get("Type", ""),
                    enc.get("Angle (°)", ""),
                    "",   # Circuit — user input
                    enc.get("Panflex L", ""),
                    enc.get("Panflex R", ""),
                    "",   # Amp ID L — user input
                    "",   # Amp ID R — user input
                    "",   # Amp Ch — user input
                ]
                for col_idx, val in enumerate(static_vals, 1):
                    c = cover.cell(row=sum_row, column=col_idx, value=val)
                    c.font = Font(name="Arial", size=9)
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    if col_idx in USER_INPUT_COLS:
                        c.fill = INPUT_FILL
                    elif col_idx <= 2:
                        c.fill = group_color
                    else:
                        c.fill = row_fill
                    c.border = thin_border()
                cover.row_dimensions[sum_row].height = 13

                # Track this cover row for back-referencing from group sheets
                _cover_enc_rows[f"{sheet_name}__{source['name']}__{enc_idx}"] = sum_row
                sum_row += 1

        sum_row += 1  # spacer between groups

    # Add dropdown + conditional formatting on Circuit col in cover
    if any(_cover_enc_rows.values()):
        all_cover_rows = list(_cover_enc_rows.values())
        min_r, max_r = min(all_cover_rows), max(all_cover_rows)
        circuit_range = f"F{min_r}:F{max_r}"
        dv_cover = DataValidation(type="list", formula1='"A,B,C,D,E,F,G,H,I,J"', allow_blank=True)
        cover.add_data_validation(dv_cover)
        dv_cover.add(circuit_range)
        for letter, hex_color in CIRCUIT_COLORS.items():
            cover.conditional_formatting.add(circuit_range, CellIsRule(
                operator="equal",
                formula=[f'"{letter}"'],
                fill=PatternFill("solid", start_color=hex_color, end_color=hex_color)
            ))

    # ── Back-fill group sheets: replace user-input cols with formulas to cover ─
    for group_name, sources in groups.items():
        sheet_name = re.sub(r'[\/*?:\[\]]', '', group_name)[:31]
        ws = wb[sheet_name]
        for source in sources:
            enclosures = source["enclosures"]
            columns    = source["columns"]
            if not enclosures:
                continue
            enc_start = _enc_start_rows.get(f"{sheet_name}__{source['name']}", None)
            if enc_start is None:
                continue

            # Map column name to cover sheet column letter
            cover_col_map = {
                "Circuit":  "F",
                "Amp ID L": "I",
                "Amp ID R": "J",
                "Amp Ch":   "K",
            }

            for enc_idx, enc in enumerate(enclosures):
                ws_row = enc_start + enc_idx
                cover_row = _cover_enc_rows.get(f"{sheet_name}__{source['name']}__{enc_idx}")
                if cover_row is None:
                    continue
                for col_idx, key in enumerate(columns, 1):
                    if key in cover_col_map:
                        cover_col = cover_col_map[key]
                        c = ws.cell(row=ws_row, column=col_idx)
                        c.value = f"='Report Info'!{cover_col}{cover_row}"
                        c.fill = INPUT_FILL
                        c.font = BODY_FONT
                        c.alignment = CENTER
                        c.border = thin_border()

    wb.save(output_path)


def write_pdf(groups, output_path, report_name="", report_date=""):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_LEFT

    doc = SimpleDocTemplate(str(output_path), pagesize=A4,
                            leftMargin=20*mm, rightMargin=20*mm,
                            topMargin=20*mm, bottomMargin=20*mm)

    NAVY  = colors.HexColor("#282C34")
    BLUE  = colors.HexColor("#3A3F4B")
    MBLUE = colors.HexColor("#282C34")
    LBLUE = colors.HexColor("#FAF3E0")
    WHITE = colors.HexColor("#FAFAFA")
    AMBER = colors.HexColor("#F5A623")

    title_style   = ParagraphStyle("t",  fontSize=11, textColor=colors.HexColor("#F5A623"), fontName="Helvetica-Bold", leading=14)
    source_style  = ParagraphStyle("s",  fontSize=9,  textColor=WHITE, fontName="Helvetica-Bold", leading=12)
    section_style = ParagraphStyle("ss", fontSize=9,  textColor=colors.white, fontName="Helvetica-Bold", leading=11)
    label_style   = ParagraphStyle("l",  fontSize=9,  textColor=NAVY,  fontName="Helvetica-Bold")
    value_style   = ParagraphStyle("v",  fontSize=9,  fontName="Helvetica")

    def banner(text, style, bg, width=170*mm):
        t = Table([[Paragraph(text, style)]], colWidths=[width])
        t.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), bg),
            ("TOPPADDING",    (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
            ("LEFTPADDING",   (0,0), (-1,-1), 8),
        ]))
        return t

    story = []

    # ── Cover block ───────────────────────────────────────────────────────────
    if report_name:
        cover_title = ParagraphStyle("ct", fontSize=10, textColor=colors.HexColor("#F5A623"),
                                     fontName="Helvetica-Bold", leading=13)
        cover_sub   = ParagraphStyle("cs", fontSize=8, textColor=colors.HexColor("#AAAAAA"),
                                     fontName="Helvetica", leading=10)
        cover_table = Table([
            [Paragraph(report_name, cover_title)],
            [Paragraph(f"Date: {report_date}" if report_date else "", cover_sub)],
        ], colWidths=[170*mm])
        cover_table.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), colors.HexColor("#282C34")),
            ("TOPPADDING",    (0,0), (-1,-1), 5),
            ("BOTTOMPADDING", (0,0), (-1,-1), 5),
            ("LEFTPADDING",   (0,0), (-1,-1), 10),
        ]))
        story.append(cover_table)
        story.append(Spacer(1, 3*mm))

    first_group = True
    for group_name, sources in groups.items():
        if not first_group:
            story.append(PageBreak())
        first_group = False
        story.append(banner(f"Group: {group_name}", title_style, NAVY))
        story.append(Spacer(1, 2*mm))
        first_source = True
        for source in sources:
            if not first_source:
                story.append(PageBreak())
            first_source = False
            story.append(banner(source["name"], source_style, BLUE))
            story.append(Spacer(1, 1*mm))
            physical = source["physical"]
            if physical:
                story.append(banner("Physical Configuration", section_style, colors.HexColor("#4A5060")))
                story.append(Spacer(1, 1*mm))
                group_label_style = ParagraphStyle("gl", fontSize=8, textColor=colors.HexColor("#2E75B6"),
                                                   fontName="Helvetica-Bold", leading=10)
                for group_label, items in grouped_physical_items(physical):
                    story.append(Paragraph(group_label, group_label_style))
                    story.append(Spacer(1, 0.5*mm))
                    rows = []
                    for i in range(0, len(items), 2):
                        lk, lv = items[i]
                        rk, rv = items[i+1] if i+1 < len(items) else ("", "")
                        rows.append([
                            Paragraph(lk, label_style), Paragraph(str(lv), value_style),
                            Paragraph(rk, label_style), Paragraph(str(rv), value_style),
                        ])
                    pt = Table(rows, colWidths=[42*mm, 40*mm, 42*mm, 46*mm])
                    ps = [("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#B0B0B0")),
                          ("TOPPADDING", (0,0), (-1,-1), 3),
                          ("BOTTOMPADDING", (0,0), (-1,-1), 3),
                          ("LEFTPADDING", (0,0), (-1,-1), 5)]
                    for i in range(len(rows)):
                        ps.append(("BACKGROUND", (0,i), (-1,i), LBLUE if i%2==0 else WHITE))
                    pt.setStyle(TableStyle(ps))
                    story.append(pt)
                    story.append(Spacer(1, 2*mm))
                story.append(Spacer(1, 1*mm))
            enclosures = source["enclosures"]
            columns    = source["columns"]
            if enclosures:
                story.append(banner("Per-Enclosure Geometry", section_style, colors.HexColor("#4A5060")))
                story.append(Spacer(1, 1*mm))
                # PDF is a rigging reference — exclude user-input columns
                pdf_exclude = {"Circuit", "Amp ID L", "Amp ID R", "Amp Ch"}
                pdf_columns = [c for c in columns if c not in pdf_exclude]
                col_w = 170*mm / len(pdf_columns)
                enc_rows = [pdf_columns] + [[str(e.get(k, "")) for k in pdf_columns] for e in enclosures]
                et = Table(enc_rows, colWidths=[col_w] * len(pdf_columns))
                es = [("BACKGROUND",    (0,0), (-1,0),  MBLUE),
                      ("TEXTCOLOR",     (0,0), (-1,0),  AMBER),
                      ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
                      ("FONTSIZE",      (0,0), (-1,-1), 8),
                      ("ALIGN",         (0,0), (-1,-1), "CENTER"),
                      ("GRID",          (0,0), (-1,-1), 0.5, colors.HexColor("#B0B0B0")),
                      ("TOPPADDING",    (0,0), (-1,-1), 3),
                      ("BOTTOMPADDING", (0,0), (-1,-1), 3)]
                CARD_COLOR = colors.HexColor("#F9DCDC")
                for i in range(1, len(enc_rows)):
                    is_cardioid = enclosures[i-1].get("Type", "").endswith("_C")
                    if is_cardioid:
                        bg = CARD_COLOR
                    else:
                        bg = LBLUE if enc_color_index(i-1, enclosures) == 0 else WHITE
                    es.append(("BACKGROUND", (0,i), (-1,i), bg))
                et.setStyle(TableStyle(es))
                story.append(et)
            story.append(Spacer(1, 2*mm))
    doc.build(story)

def process_pdf(pdf_path):
    print(f"  Processing: {pdf_path.name}")
    text   = extract_text(pdf_path)
    groups = parse_document(text)
    name, date = extract_metadata(text)
    total  = sum(len(s) for s in groups.values())
    print(f"  Found {len(groups)} group(s), {total} source(s)")
    xlsx_path  = OUTPUT_DIR / (pdf_path.stem + ".xlsx")
    pdf_path2  = OUTPUT_DIR / (pdf_path.stem + "_report.pdf")
    write_excel(groups, xlsx_path, report_name=name, report_date=date)
    write_pdf(groups, pdf_path2, report_name=name, report_date=date)
    print(f"  Saved: {xlsx_path.name}")
    print(f"  Saved: {pdf_path2.name}")

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    if len(sys.argv) > 1:
        pdf_path = DATA_DIR / sys.argv[1]
        if not pdf_path.exists():
            print(f"Error: {pdf_path} not found.")
            sys.exit(1)
        pdfs = [pdf_path]
    else:
        pdfs = sorted(DATA_DIR.glob("*.pdf"))
        if not pdfs:
            print(f"No PDF files found in {DATA_DIR}")
            sys.exit(0)
    print(f"Found {len(pdfs)} PDF(s) to process.\n")
    for pdf in pdfs:
        process_pdf(pdf)
    print("\nDone.")

if __name__ == "__main__":
    main()
